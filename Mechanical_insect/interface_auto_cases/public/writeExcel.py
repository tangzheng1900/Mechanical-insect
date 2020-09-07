# -*- coding: utf-8 -*-
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : writeExcel.py
# @Software: PyCharm


from openpyxl import Workbook, load_workbook
import interface_auto_cases.conf.NacosConfig as NC
from interface_auto_cases.public.logger import Log
import interface_auto_cases.public.get_mysql_info  as gmi

logger = Log('writeExcel', NC.log_path)


class writMysql:
    def __init__(self):
        logger.info("开始写入数据库用例测试结果！")

    def write_Mysql(self, i, dict):
        casedata = gmi.getMysqlInfo(NC.config).auto_testcase(self.version)


class writeExcel:
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    def write_Excel(self, i, dict):
        # 写入用例返回结果
        try:
            wb_new = load_workbook(self.file)
            sheet = wb_new[self.sheet]
            sheet.cell(i, 8).value = dict['code']
            # sheet.cell(i,9).value = dcit['sql_result']
            sheet.cell(i, 10).value = dict['result']
            if 'data' in dict.keys():
                sheet.cell(i, 11).value = str(dict['data'])
            wb_new.save(self.file)
            logger.info('执行写入excel成功！')
        except Exception as e:
            logger.info('数据写入失败！%s' % e)
            raise e

    def Excel_dellog(self):
        # 清除原始记录
        wb_new = load_workbook(self.file)
        sheet = wb_new[self.sheet]
        for i in range(sheet.max_row - 1):
            sheet.cell(row=i + 2, column=8).value = ""
            sheet.cell(row=i + 2, column=9).value = ""
            sheet.cell(row=i + 2, column=10).value = ""
            sheet.cell(row=i + 2, column=11).value = ""
        wb_new.save(self.file)
        logger.info('清除excel原始记录成功！')


def writeexcel(dict):
    file = NC.test_data_path
    sheet1 = 'Sheet2'
    try:
        wb_new = load_workbook(file)
        sheet = wb_new[sheet1]
        i = sheet.max_row + 1
        if 'restult' in dict.keys():
            sheet.cell(2, 1).value = dict['restult']
            logger.info('执行excel写入统计测试记录成功！')
        else:
            sheet.cell(i, 2).value = dict['sum']
            sheet.cell(i, 3).value = dict['ok']
            sheet.cell(i, 4).value = dict['fail']
            sheet.cell(i, 5).value = dict['error']
            sheet.cell(i, 6).value = dict['error_1']
            logger.info('执行excel写入数据测试记录成功！')
        wb_new.save(file)
    except Exception as e:
        logger.info('执行excel写入测试记录失败！%s' % e)
        raise e


if __name__ == '__main__':
    writeexcel({'sum': 88, 'ok': 99, 'fail': 66, 'error': 77})
    header = ['学号', '姓名', '性别', '班级']
    data = [['001', 'sunny', 'F', 'python5'], ['002', '小胖雨', 'M', 'python5']]
    print(writeExcel(NC.test_data_path, 'Sheet1').write_Excel(1, {'code': 'zgh', 'sql_result': 000, 'result': 333,
                                                                  'msg': 555}))
