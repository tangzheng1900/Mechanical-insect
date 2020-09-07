# -*- coding: utf-8 -*-
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : readExcel.py
# @Software: PyCharm


import os
from interface_auto_cases.conf import Allpath
from interface_auto_cases.public.config import config
from openpyxl import load_workbook
import interface_auto_cases.conf.NacosConfig as NC
import interface_auto_cases.public.get_mysql_info  as gmi
from interface_auto_cases.public.logger import Log

loger = Log('readExcel', NC.log_path)
url_1 = str(gmi.project_url)
mode = config().read_config(Allpath.case_conf_path, 'FLAG', 'mode')
loger.info("用例执行模式（0全部执行，1指定接口，2区间接口）：%s" % mode)
case_list = config().read_config(Allpath.case_conf_path, 'FLAG', 'case_list')
case_list1 = config().read_config(Allpath.case_conf_path, 'FLAG', 'case_list1')
loger.info("配置用例列表：指定{},区间{}".format(case_list, case_list1))


class Case_Id:
    a, b = case_list1[0], case_list1[1]
    c = []
    if a > 0 or b > 0:
        if a <= b:
            for i in range(a, b + 1):
                c.append(i)
    if len(c) > 0:
        case_list1 = c
        loger.info("当前使用临时配置用例列表：%s" % case_list)
    else:
        case_list1 = case_list


class Mysql_Casedata:
    def __init__(self, version):
        self.version = version
        loger.info("开始处理数据库用例自动化格式！")

    def mysql_Casedata(self, mode=0):
        data_list = []
        casedata = gmi.getMysqlInfo(NC.config).auto_testcase(self.version)
        if mode == 1:
            loger.info("------------版本指定用例集------------")
            pass
        elif mode == 2:
            loger.info("------------版本区间用例集------------")
            pass
        else:
            loger.info("------------版本全部用例集------------")
            for i in range(len(casedata)):
                id = casedata[i]["id"]
                method = casedata[i]["method"]
                cases_name = casedata[i]["cases_name"]
                url = url_1 + casedata[i]["url"]
                data = eval(casedata[i]["data"])
                sql = eval(casedata[i]["sql"])
                code = casedata[i]["code"]
                dict = {'id': id, 'case_name': cases_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                        'code': code}
                data_list.append(dict)

        return data_list


class readExcel:
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    def read_Excel(self):
        rd = load_workbook(self.file)
        sheet = rd[self.sheet]
        loger.info('获取表单成功：%s' % sheet)

        # initphone=self.get_init_phone()
        # print(initphone)#当作一个局部变量来看这个，一轮循环下来后在更新到excel
        # 0全部，1指点，2指定区间
        data_list = []
        if mode == 0:
            for i in range(sheet.max_row - 1):
                method = sheet.cell(row=i + 2, column=2).value
                case_name = sheet.cell(row=i + 2, column=3).value
                url = url_1 + sheet.cell(row=i + 2, column=4).value
                data = eval(sheet.cell(row=i + 2, column=5).value)  # 取值excel中的字典
                sql = eval(sheet.cell(row=i + 2, column=6).value)
                code = sheet.cell(row=i + 2, column=7).value
                dict = {'id': i + 1, 'case_name': case_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                        'code': code}
                data_list.append(dict)
            # self.update_phone(initphone)#把最后的值更新到excel中

        if mode == 1:
            for i in range(sheet.max_row - 1):
                id = sheet.cell(row=i + 2, column=1).value
                if id in case_list:
                    # print(i)
                    method = sheet.cell(row=i + 2, column=2).value
                    case_name = sheet.cell(row=i + 2, column=3).value
                    url = url_1 + str(sheet.cell(row=i + 2, column=4).value)
                    data = eval(sheet.cell(row=i + 2, column=5).value)  # 取值excel中的字典
                    sql = eval(sheet.cell(row=i + 2, column=6).value)
                    code = sheet.cell(row=i + 2, column=7).value
                    dict = {'id': i + 1, 'case_name': case_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                            'code': code}
                    data_list.append(dict)
        if mode == 2:
            for i in range(sheet.max_row - 1):
                id = sheet.cell(row=i + 2, column=1).value
                if id in Case_Id.case_list1:
                    # print(i)
                    method = sheet.cell(row=i + 2, column=2).value
                    case_name = sheet.cell(row=i + 2, column=3).value
                    url = url_1 + str(sheet.cell(row=i + 2, column=4).value)
                    data = eval(sheet.cell(row=i + 2, column=5).value)  # 取值excel中的字典
                    sql = eval(sheet.cell(row=i + 2, column=6).value)
                    code = sheet.cell(row=i + 2, column=7).value
                    dict = {'id': i + 1, 'case_name': case_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                            'code': code}
                    data_list.append(dict)

        # print(data_list)
        return data_list


def readexcel():
    file = Allpath.test_data_path
    sheet = 'Sheet2'
    rd = load_workbook(file)
    sheet = rd[sheet]
    loger.info('获取历史测试表单成功：%s' % sheet)
    dict = {'restult': '', 'sum': [], 'ok': [], 'fail': [], 'error': [], 'error_1': []}
    max = sheet.max_row + 1
    mix = sheet.max_row - 10
    if mix <= 1:
        mix = 2
    else:
        mix = max - 10
    for i in range(mix, max):
        dict['restult'] = sheet.cell(row=2, column=1).value
        dict['sum'].append(sheet.cell(row=i, column=2).value)
        dict['ok'].append(sheet.cell(row=i, column=3).value)
        dict['fail'].append(sheet.cell(row=i, column=4).value)
        dict['error'].append(sheet.cell(row=i, column=5).value)
        dict['error_1'].append(sheet.cell(row=i, column=5).value)
    return dict


if __name__ == '__main__':
    Mysql_Casedata(gmi.version).mysql_Casedata()
